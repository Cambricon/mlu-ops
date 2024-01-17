#!/usr/bin/env python3
#coding:utf-8

import pandas as pd
import subprocess
import re
import os
import google.protobuf.text_format
import hashlib
import numpy as np
import xlwt
import xlrd
import xlutils
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pandas import ExcelWriter
import argparse
import collections
#import utils
#import time,sys

proto_files = []

def check_mluop_proto(parser):
    # if mlu_op_test_pb2.py do not exist, generate it
    if not os.path.exists("mlu_op_test_pb2.py"):
        proto_dir = os.path.abspath(
            os.path.realpath(__file__) +
            "/../../../bangc-ops/test/mlu_op_gtest/pb_gtest/mlu_op_test_proto")
        if not os.path.exists(proto_dir + "/mluop_test.proto"):
            print('{} do not exist, please check!'.format(proto_dir +
                                                          "/mluop_test.proto"))
            parser.print_help()
            exit()

        cwd = os.path.abspath(os.path.realpath(__file__) + "/../")
        cmd_args = [
            "protoc", "--python_out", cwd, "--proto_path", proto_dir,
            proto_dir + "/mluop_test.proto"
        ]
        try:
            cmd_ret = subprocess.run(cmd_args)
        except Exception as e:
            print('{} run failed, please check!'.format(' '.join(cmd_args)))
            exit()


def gci(path):
    for root,dirs,files in os.walk(path):
        for filespath in files:
            fi_d = os.path.join(os.path.abspath(root), filespath)
            if fi_d.endswith(".pb") or fi_d.endswith(".prototxt"):
                proto_files.append(fi_d)
    return proto_files


def resolve_case(path):
    node = mlu_op_test_pb2.Node()
    if path.endswith(".prototxt"):
        with open(path) as f:
            google.protobuf.text_format.Parse(f.read(), node)
    elif path.endswith(".pb"):
        with open(path, "rb") as f:
            node.ParseFromString(f.read())
    return get_node_info(node)


def get_node_info(node):
    # Init
    result = collections.OrderedDict()
    try:
        input_dtype = []
        input_layout = []
        inputs_shape = []
        input_stride = []
        onchip_dtype = []
        op_name = node.op_name
        result["op_name"] = op_name
        for one_input in node.input:
            # do not need HasField of hasattr
            inputs_shape.append([str(i) for i in list(one_input.shape.dims)])
            input_stride.append(
                [str(i) for i in list(one_input.shape.dim_stride)])
            input_layout.append(
                mlu_op_test_pb2.TensorLayout.Name(one_input.layout))
            input_dtype.append(mlu_op_test_pb2.DataType.Name(one_input.dtype))
            # onchip_dtype and params should be carefully
            if one_input.HasField('onchip_dtype'):
                onchip_dtype.append(
                    mlu_op_test_pb2.DataType.Name(one_input.onchip_dtype))
        result["inputs"] = ';'.join(
            [','.join(shape) for shape in inputs_shape])
        result["input_stride"] = ';'.join(
            [','.join(strid) for strid in input_stride])
        result["input_datatype"] = ';'.join(input_dtype)
        result["input_layout"] = ';'.join(input_layout)
        output_stride = []
        outputs_shape = []
        output_dtype = []
        output_layout = []
        for one_output in node.output:
            # do not need HasField
            outputs_shape.append([str(i) for i in list(one_output.shape.dims)])
            output_stride.append(
                [str(i) for i in list(one_output.shape.dim_stride)])
            output_layout.append(
                mlu_op_test_pb2.TensorLayout.Name(one_output.layout))
            output_dtype.append(mlu_op_test_pb2.DataType.Name(one_output.dtype))
            if one_output.HasField('onchip_dtype'):
                onchip_dtype.append(
                    mlu_op_test_pb2.DataType.Name(one_output.onchip_dtype))
        result["outputs"] = ';'.join(
            [','.join(shape) for shape in outputs_shape])
        result["output_stride"] = ';'.join(
            [','.join(stride) for stride in output_stride])
        result["output_datatype"] = ';'.join(output_dtype)
        result["output_layout"] = ';'.join(output_layout)
        params = ""
        result["onchip_datatype"] = ';'.join(onchip_dtype)
        # double check
        if hasattr(node, op_name + "_param"):
            if node.HasField(op_name + "_param"):
                # TO DO: handle activation_param
                params += str(getattr(node,
                                      op_name + "_param")).replace('\n', ';')
        result["params"] = params
    except:
        pass
    return result


def puts_modify(flag_diff, flag, search_columns_layout, columns_list, layout):
    seq_num = re.findall("\d+", search_columns_layout)
    layout_type = layout.split('_')
    replace_dim = layout_type[0] + 's' + str(seq_num[0])
    temp = flag_diff.split('_')
    flag_str1 = temp[1]
    for search_columns_dims in columns_list:
        if replace_dim in search_columns_dims:
            str_list = search_columns_dims.split('_')
            search_columns_dims_re = str_list[0] + "_" + flag_str1
            df.rename(columns={search_columns_dims: search_columns_dims_re},
                      inplace=True)


def search_columns_layout_Uncommon1(layout, df):
    layout_flag = {
        'flag_NCHW': 0,
        ' flag_NHWC': 0,
        ' flag_WHCN': 0,
        ' flag_NDHWC': 0,
        ' flag_HWCN': 0,
        'flag_NCDHW': 0,
        'flag_TNC': 0,
        'flag_NTC': 0
    }
    flag = df.shape[0]
    columns_list = df.columns.tolist()
    for search_columns_layout in columns_list:
        if layout in search_columns_layout:
            for index, row in df.iterrows():
                if "ARRAY" in row[search_columns_layout]:
                    break
                if row[search_columns_layout] == " ":
                    for key in layout_flag:
                        layout_flag[key] += 1
                else:
                    for key in layout_flag:
                        keys = key.split('_')
                        if keys[1] in row[search_columns_layout]:
                            layout_flag[key] += 1
            for key in layout_flag:
                if layout_flag[key] == flag:
                    puts_modify(key, flag, search_columns_layout, columns_list,
                                layout)
            for key in layout_flag:
                layout_flag[key] = 0


def search_columns_layout_Uncommon(layout, df):
    flag_NCHW, flag_NHWC, flag_WHCN, flag_NDHWC, flag_HWCN, flag_NCDHW, flag_TNC, flag_NTC = 0, 0, 0, 0, 0, 0, 0, 0
    flag = df.shape[0]
    columns_list = df.columns.tolist()
    for search_columns_layout in columns_list:
        if layout in search_columns_layout:
            for index, row in df.iterrows():
                if "NCHW" in row[search_columns_layout]:
                    flag_NCHW += 1
                elif "NHWC" in row[search_columns_layout]:
                    flag_NHWC += 1
                elif "WHCN" in row[search_columns_layout]:
                    flag_WHCN += 1
                elif "HWCN" in row[search_columns_layout]:
                    flag_NWCN += 1
                elif "NDHWC" in row[search_columns_layout]:
                    flag_NDHWC += 1
                elif "TNC" in row[search_columns_layout]:
                    flag_TNC += 1
                elif "NCDHW" in row[search_columns_layout]:
                    flag_NCDHW += 1
                elif "NTC" in row[search_columns_layout]:
                    flag_NTC += 1
                elif row[search_columns_layout] == " ":
                    flag_NCHW += 1
                    flag_NHWC += 1
                    flag_HWCN += 1
                    flag_WHCN += 1
                    flag_NDHWC += 1
                    flag_TNC += 1
                    flag_NCDHW += 1
                    flag_NTC += 1
                elif "ARRAY" in row[search_columns_layout]:
                    break
            if flag_NCHW == flag:
                puts_modify("flag_NCHW", flag, search_columns_layout,
                            columns_list, layout)
            elif flag_NHWC == flag:
                puts_modify("flag_NHWC", flag, search_columns_layout,
                            columns_list, layout)
            elif flag_WHCN == flag:
                puts_modify("flag_WHCN", flag, search_columns_layout,
                            columns_list, layout)
            elif flag_HWCN == flag:
                puts_modify("flag_HWCN", flag, search_columns_layout,
                            columns_list, layout)
            elif flag_NDHWC == flag:
                puts_modify("flag_NDHWC", flag, search_columns_layout,
                            columns_list, layout)
            elif flag_TNC == flag:
                puts_modify("flag_TNC", flag, search_columns_layout,
                            columns_list, layout)
            elif flag_NCDHW == flag:
                puts_modify("flag_NCDHW", flag, search_columns_layout,
                            columns_list, layout)
            elif flag_NTC == flag:
                puts_modify("flag_NTC", flag, search_columns_layout,
                            columns_list, layout)
            else:
                break
            flag_NCHW, flag_NHWC, flag_HWCN, flag_NDHWC, flag_NWCN = 0, 0, 0, 0, 0


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--case_path',
                        type=str,
                        help='path of the input file, can be prototxt or pb ',
                        required=True)
    parser.add_argument(
        '--xlsx_path',
        type=str,
        help=
        'path of the output excel, the default file name is case_path.xlsx, i.e. op_tensor.xlsx',
        required=False)
    opt, unknown = parser.parse_known_args()
    check_mluop_proto(parser)
    import mlu_op_test_pb2
    logpath = os.path.abspath(opt.case_path)
    proto_files = gci(logpath)
    if len(proto_files) == 0:
        print("the folder is empty,please check {}!".format(logpath))
        exit()
    if opt.xlsx_path != None:
        if not opt.xlsx_path.endswith('xlsx'):
            opt.xlsx_path = opt.xlsx_path + '.xlsx'
    else:
        opt.xlsx_path = logpath.split('/')[-1] + '.xlsx'

    print("parsing {}...".format(os.path.abspath(opt.case_path)))
    book = xlwt.Workbook(encoding='utf-8', style_compression=0)
    sheet = book.add_sheet('analyse', cell_overwrite_ok=True)
    cols = ('op_name', 'inputs', 'input_stride', 'input_datatype',
            'input_layout', 'outputs', 'output_stride', 'output_datatype',
            'output_layout', 'onchip_dtype', 'params', 'path')
    for i in range(0, 12):
        sheet.write(0, i, cols[i])
    num = 1
    for seq in proto_files:
        path = os.path.join(logpath, seq)
        result = resolve_case(path)
        lst = []
        # dict data to list data
        for key in result:
            nums = [key, result[key]]
            lst.append(nums)
        #list data to excel
        for i in range(0, 11):
            data = lst[i]
            sheet.write(num, i, data[1])
        sheet.write(num, 11, path)
        num += 1
    print("waiting ........")
    xlsx_path = os.path.abspath(opt.xlsx_path)
    book.save(xlsx_path)
    df = pd.read_excel(xlsx_path)
    for strs in cols[1:11]:
        df = df.fillna(0)
        df2 = df[strs].astype(str).str.split(';', expand=True)
        tmp_column = df2.shape[1]
        for col in range(0, tmp_column):
            str2 = strs + str(col + 1)
            df2.rename(columns={col: str2}, inplace=True)
        df3 = df.drop([strs], axis=1).join(df2)
        df = df3
    search_columns_layout_Uncommon("input_layout", df)
    search_columns_layout_Uncommon("output_layout", df)
    columns_length = df.shape[1]
    df.drop(df.columns[columns_length - 1], axis=1, inplace=True)
    move_path = df.pop('path')
    df.insert(columns_length - 2, 'path', move_path)
    #    for search_columns_stride in columns_list:
    #       if "stride" in search_columns_stride:
    df.replace('', str(0.0), inplace=True)
    df.to_excel(xlsx_path, index=False)
    #change excel format
    wb = load_workbook(xlsx_path)
    ws = wb[wb.sheetnames[0]]
    width1 = 2.0
    height = width1 * (2.2862 / 0.3612) * 1.5
    column_widths = (
        df.columns.to_series().apply(lambda x: len(x.encode('gbk'))).values)
    max_widths = (df.astype(str).applymap(lambda x: len(x.encode('gbk'))).agg(
        max).values)
    widths = np.max([column_widths, max_widths], axis=0)
    for i in range(0, ws.max_row):
        ws.row_dimensions[i].height = height
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width + 2
    wb.save(xlsx_path)
    print("parse succeed,saved!")
